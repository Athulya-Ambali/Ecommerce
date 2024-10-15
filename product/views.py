from django.shortcuts import render,redirect,get_object_or_404
from django.contrib.auth import login,logout
from django.conf import settings
from product.models import UserA,Product,Order
from django.views import View
from product.forms import ProductForm,ExcelForm,OrderForm
from openpyxl import Workbook,load_workbook
from django.http import HttpResponse
from django.http import JsonResponse

# Create your views here.


class UserLogin(View):
    def get(self, request):
        return render(request,'login.html')
    
    def post(self,request):
        if request.method == 'POST':
            uname=request.POST.get('uname')
            password=request.POST.get('password')
            print(uname,password)
            
            user=UserA.objects.get(uname=uname,password=password)
            request.session['user']=user.id
            
            if user is not None:
                login(request,user)
                return redirect('home')
            return render(request,'login.html', {'error': 'Invalid credentials'})
        

class UserLogout(View):
    def get(self,request):
        logout(request)
        return redirect('login')


class Home(View):
    def get(self,request):
        return render(request, 'home.html')

            
class ProductCreate(View):
    def get(self,request):
        user1=UserA.objects.get(id=request.session.get('user'))
        if user1.uname in settings.ALLOWED_USERS:
            form=ProductForm
            return render(request, 'product_create.html', {'form': form})
        else:
            return redirect('product_list')
    
    
    def post(self,request):
         user1=UserA.objects.get(id=request.session.get('user'))
         if user1.uname in settings.ALLOWED_USERS:
             form=ProductForm(request.POST)
             if form.is_valid():
                 form.save()
                 return redirect('home')
             return render(request, 'product_create.html', {'form': form})
         else:
            return redirect('product_list')
    

class ProductList(View):
    def get(self,request):
        products=Product.objects.all()
        return render(request, 'product_list.html', {'products': products,'allowed_users':allowed_users})

    def post(self, request):
        product_id = request.POST.get('product_id')
        stock_to_add = request.POST.get('stock_to_add')

        if product_id and stock_to_add and stock_to_add.isdigit():
            product_instance = get_object_or_404(Product, id=product_id)
            product_instance.stock += int(stock_to_add)
            product_instance.save()

        return redirect('product_list')
        
class ProductUpdate(View):
    def get(self,request, id):
        product_instance=Product.objects.get(id=id)
        form=ProductForm(instance=product_instance)
        return render(request, 'product_update.html', {'form': form})
    
    def post(self,request, id):
        product_instance=Product.objects.get(id=id)
        form=ProductForm(request.POST, instance=product_instance)
        if form.is_valid():
            form.save()
            return redirect('product_list')
        return render(request, 'product_update.html', {'form': form})
    
class ProductDelete(View):
    def get(self,request, id):
        product_instance=Product.objects.get(id=id)
        product_instance.delete()
        return redirect('product_list')
    



class ExportProducts(View):

    def get(self, request, *args, **kwargs):
       
        wb = Workbook()
        ws = wb.active
        ws.title = "Product Data"
        
      
        headers = ['ID', 'Name', 'Price', 'Stock', 'Status']
        ws.append(headers)  
        
        products = Product.objects.all()
        for product in products:
            ws.append([product.id, product.name, product.price, product.stock, product.get_status_display()])
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="products.xlsx"'
        
        wb.save(response)
        return response


  

class ImportProducts(View):
    def get(self, request, *args, **kwargs):
        form = ExcelForm()  
        return render(request, 'import_products.html', {'form': form})

    def post(self, request, *args, **kwargs):
        form = ExcelForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            wb = load_workbook(file)
            ws = wb.active
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                try:
                    
                    product_id, name, price, stock, status = row[:5]  

                    product, created = Product.objects.update_or_create(
                        id=product_id,
                        defaults={
                            'name': name,
                            'price': price,
                            'stock': stock,
                            'status': status if status in dict(Product.STATUS_CHOICES) else Product.AVAILABLE
                        }
                    )
                except ValueError as e:
                    print(f"Skipping row: {row}. Error: {e}")

            return redirect('product_list')  
        return render(request, 'import_products.html', {'form': form})
    


class OrderCreate(View):
    def get(self, request):
        product_id = request.GET.get('product_id', None)
        form = OrderForm()
        if product_id:
            product_instance = get_object_or_404(Product, id=product_id)
            form = OrderForm(initial={'product': product_instance})
        
        return render(request, 'order_form.html', {'form': form})
    
    def post(self, request):
        form = OrderForm(request.POST)
        if form.is_valid():
            order = form.save(commit=False)
            user1=UserA.objects.get(id=request.session.get('user'))
            order.user = user1
            print(order.user)
            order.save()
            return redirect('order_list')
        return render(request, 'order_form.html', {'form': form})



class OrderList(View):
    def get(self, request):
        orders = Order.objects.filter(user=request.session.get('user'))  # Filter orders for the current user
        return render(request, 'order_list.html', {'orders': orders})



class OrderUpdate(View):
    def get(self, request, id):
        order = get_object_or_404(Order, id=id, user=request.user)  # Ensure only the user can access their own orders
        form = OrderForm(instance=order)
        return render(request, 'order_update.html', {'form': form, 'order': order})

    def post(self, request, id):
        order = get_object_or_404(Order, id=id, user=request.user)
        form = OrderForm(request.POST, instance=order)
        if form.is_valid():
            form.save()
            return redirect('order_list')
        return render(request, 'order_update.html', {'form': form, 'order': order})


class OrderDelete(View):
    def get(self, request, id):
        order = get_object_or_404(Order, id=id, user=request.user)  # Ensure the user can only delete their own orders
        product = get_object_or_404(Product, name=order.product)
        product.stock += order.quantity
        product.save()
        order.delete()
        return redirect('order_list')


class OrderDetailView(View):
    def get(self, request, id):
        order = get_object_or_404(Order, id=id, user=request.user)
        return render(request, 'order_detail.html', {'order': order})
    



# class OrderCreate(View):
#     def get(self, request):
#         product_id = request.GET.get('product_id', None)
#         form = OrderForm()
#         if product_id:
#             product_instance = get_object_or_404(Product, id=product_id)
#             form = OrderForm(initial={'product': product_instance})
        
#         return render(request, 'order_form.html', {'form': form})
    
#     def post(self, request):
#         form = OrderForm(request.POST)
#         if form.is_valid():
#             form.save()
#             return redirect('order_list')
#         return render(request, 'order_form.html', {'form': form})


# class OrderList(View):
#     def get(self,request):
#         orders=Order.objects.all()
#         return render(request, 'order_list.html', {'orders': orders})

# class OrderDetailView(View):
#     def get(self, request, id):
#         order = get_object_or_404(Order, id=id)
#         return render(request, 'order_detail.html', {'order': order})

   
# class OrderUpdate(View):
#     def get(self, request, id):
#         order = get_object_or_404(Order, id=id)
#         form = OrderForm(instance=order)
#         return render(request, 'order_update.html', {'form': form, 'order': order})

#     def post(self, request, id):
#         order = get_object_or_404(Order, id=id)
#         form = OrderForm(request.POST, instance=order)
#         if form.is_valid():
#             form.save()
#             return redirect('order_list') 
#         return render(request, 'order_update.html', {'form': form, 'order': order})


# class OrderDelete(View):
#     def get(self, request, id):
#         order = get_object_or_404(Order, id=id)
#         product = get_object_or_404(Product, name=order.product)
#         product.stock += order.quantity
#         product.save()
#         order.delete()
#         return redirect('order_list')




# class ProductPriceView(View):
#     def get(self, request, product_id):
#         product = get_object_or_404(Product, id=product_id)
#         return JsonResponse({'price': product.price})







